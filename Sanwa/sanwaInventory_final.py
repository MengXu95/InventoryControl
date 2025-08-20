#!/usr/bin/env python3
"""
Inventory Management CLI Tool
A command-line tool for processing inventory data and generating ordering recommendations.

Usage:
    python inventory_cli.py [options]
    inventory-manager [options]  # if installed as package
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
import re
import os
import glob
import argparse
import sys
from pathlib import Path

# Version info
__version__ = "1.0.0"
__author__ = "Inventory Management System"


class InventoryManager:
    """Main inventory management class"""

    def __init__(self, verbose=False):
        self.verbose = verbose
        self.log("Inventory Management System initialized")

    def log(self, message, level="INFO"):
        """Simple logging function"""
        if self.verbose or level in ["ERROR", "SUCCESS"]:
            timestamp = datetime.now().strftime("%H:%M:%S")
            prefix = "✅" if level == "SUCCESS" else "❌" if level == "ERROR" else "ℹ️"
            print(f"[{timestamp}] {prefix} {message}")

    def find_stock_on_hand_column(self, df):
        """Dynamically find the stock on hand column regardless of date"""
        stock_patterns = [
            r'.*stock\s+on\s+hand.*',
            r'.*stock.*\d{1,2}/\d{1,2}/\d{4}.*',
            r'.*on\s+hand.*\d{1,2}/\d{1,2}/\d{4}.*'
        ]

        for col in df.columns:
            col_lower = str(col).lower()
            for pattern in stock_patterns:
                if re.search(pattern, col_lower, re.IGNORECASE):
                    return col
        return None

    def find_po_balance_column(self, df):
        """Dynamically find the PO balance column regardless of date"""
        po_patterns = [
            r'.*po\s+bal.*',
            r'.*system\s+po.*',
            r'.*purchase\s+order.*balance.*'
        ]

        for col in df.columns:
            col_lower = str(col).lower()
            for pattern in po_patterns:
                if re.search(pattern, col_lower, re.IGNORECASE):
                    return col
        return None

    def find_input_file(self, directory="."):
        """Find inventory Excel file in specified directory"""
        patterns = [
            'inventory*.xlsx',
            '*inventory*.xlsx',
            'stock*.xlsx',
            '*stock*.xlsx'
        ]

        search_dir = Path(directory)
        for pattern in patterns:
            files = list(search_dir.glob(pattern))
            if files:
                return str(files[0])

        # If no specific pattern matches, look for any Excel file
        excel_files = list(search_dir.glob('*.xlsx'))
        if excel_files:
            self.log(f"Found Excel files: {[f.name for f in excel_files]}")
            return str(excel_files[0])

        return None

    def generate_output_filename(self, input_file=None, output_dir="."):
        """Generate output filename based on current timestamp"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if input_file:
            base_name = Path(input_file).stem
            filename = f"{base_name}_ordering_recommendations_{timestamp}.xlsx"
        else:
            filename = f"ordering_recommendations_{timestamp}.xlsx"

        return str(Path(output_dir) / filename)

    def safe_int_convert(self, value):
        """Safely convert a value to integer, handling NaN and negative values"""
        if pd.isna(value):
            return 0
        try:
            result = int(float(value))
            return max(0, result)
        except (ValueError, TypeError):
            return 0

    def parse_leadtime(self, leadtime_str):
        """Parse leadtime string and return weeks as integer"""
        if pd.isna(leadtime_str):
            return 8

        leadtime_str = str(leadtime_str).strip()

        # Handle months first (convert to weeks)
        month_pattern = r'(\d+)\s*mth'
        month_match = re.search(month_pattern, leadtime_str, re.IGNORECASE)
        if month_match:
            months = int(month_match.group(1))
            return months * 4

        # Look for week patterns
        week_range_pattern = r'(\d+)\s*[-–]\s*(\d+)\s*wks?'
        match = re.search(week_range_pattern, leadtime_str, re.IGNORECASE)
        if match:
            start_weeks = int(match.group(1))
            end_weeks = int(match.group(2))
            return max(start_weeks, end_weeks)

        single_week_pattern = r'(\d+)\s*wks?'
        match = re.search(single_week_pattern, leadtime_str, re.IGNORECASE)
        if match:
            return int(match.group(1))

        all_numbers = re.findall(r'\d+', leadtime_str)
        if all_numbers:
            leadtime_candidates = []
            for num_str in all_numbers:
                num = int(num_str)
                num_pos = leadtime_str.find(num_str)
                context = leadtime_str[max(0, num_pos - 10):num_pos + len(num_str) + 10].upper()

                if any(indicator in context for indicator in ['MOQ', 'KG', 'MT']):
                    continue

                if 1 <= num <= 52:
                    leadtime_candidates.append(num)

            if leadtime_candidates:
                return max(leadtime_candidates)

        return 8

    def extract_moq(self, leadtime_str):
        """Extract MOQ from leadtime string"""
        if pd.isna(leadtime_str):
            return 0

        leadtime_str = str(leadtime_str).strip()

        # MOQ with MT/mt patterns
        moq_mt_patterns = [
            r'MOQ\s*:?\s*(\d+(?:,\d+)*(?:\.\d+)?)\s*MT',
            r'/MOQ(\d+(?:,\d+)*(?:\.\d+)?)MT',
            r'MOQ(\d+(?:,\d+)*(?:\.\d+)?)MT'
        ]

        for pattern in moq_mt_patterns:
            match = re.search(pattern, leadtime_str, re.IGNORECASE)
            if match:
                number_str = match.group(1).replace(',', '')
                return int(float(number_str) * 1000)

        # MOQ with kg patterns
        moq_kg_patterns = [
            r'MOQ\s*:?\s*(\d+(?:,\d+)*(?:\.\d+)?)\s*kg',
            r',MOQ(\d+(?:,\d+)*(?:\.\d+)?)kg',
            r'/MOQ(\d+(?:,\d+)*(?:\.\d+)?)kg'
        ]

        for pattern in moq_kg_patterns:
            match = re.search(pattern, leadtime_str, re.IGNORECASE)
            if match:
                number_str = match.group(1).replace(',', '')
                return int(float(number_str))

        # MOQ with numbers only
        moq_number_patterns = [
            r'MOQ\s*:?\s*(\d+(?:,\d+)*)',
            r'MOQ(\d+(?:,\d+)*)'
        ]

        for pattern in moq_number_patterns:
            match = re.search(pattern, leadtime_str, re.IGNORECASE)
            if match:
                number_str = match.group(1).replace(',', '')
                number = int(number_str)
                if number <= 10:
                    return number * 1000
                else:
                    return number

        return 0

    def calculate_safety_stock(self, demands, leadtime_weeks):
        """Calculate safety stock based on demand variability and leadtime"""
        if len(demands) == 0:
            return 0

        clean_demands = [self.safe_int_convert(d) for d in demands if self.safe_int_convert(d) > 0]
        if len(clean_demands) == 0:
            return 0

        demand_std = np.std(clean_demands) if len(clean_demands) > 1 else np.mean(clean_demands) * 0.2
        safety_stock = 1.65 * demand_std * np.sqrt(leadtime_weeks / 4.33)

        return int(max(safety_stock, 0))

    def inventory_ordering_policy(self, row, month_columns, stock_col, po_col):
        """Implement inventory ordering policy for a single item"""
        leadtime_weeks = self.parse_leadtime(row.get('Revised Leadtime'))
        stock_on_hand = self.safe_int_convert(row.get(stock_col, 0))
        po_balance = self.safe_int_convert(row.get(po_col, 0))

        monthly_demands = [self.safe_int_convert(row.get(month, 0)) for month in month_columns]
        safety_stock = self.calculate_safety_stock(monthly_demands, leadtime_weeks)

        num_months = len(month_columns)
        leadtime_months = max(1, round(leadtime_weeks / 4.33))

        leadtime_info = str(row.get('Revised Leadtime'))
        moq = self.extract_moq(leadtime_info)

        monthly_orders = [0.0] * num_months
        monthly_deliveries = [0.0] * num_months
        monthly_inventory_levels = [0.0] * num_months

        current_inventory = stock_on_hand

        if po_balance > 0:
            monthly_deliveries[0] += po_balance

        for month_idx in range(num_months):
            current_inventory += monthly_deliveries[month_idx]

            future_demand = 0
            look_ahead_months = min(leadtime_months, num_months - month_idx)
            for j in range(look_ahead_months):
                if month_idx + j < num_months:
                    future_demand += monthly_demands[month_idx + j]
                    if j > 0:
                        future_demand -= monthly_deliveries[month_idx + j]

            if future_demand < 0:
                future_demand = 0

            reorder_point = future_demand + safety_stock
            projected_inventory_after_demand = current_inventory - monthly_demands[month_idx]

            if projected_inventory_after_demand < reorder_point:
                order_qty = reorder_point - projected_inventory_after_demand

                if moq > 0 and order_qty > 0:
                    order_qty = max(order_qty, moq)

                monthly_orders[month_idx] = order_qty

                delivery_month = month_idx + leadtime_months
                if delivery_month < num_months:
                    monthly_deliveries[delivery_month] += order_qty

            current_inventory = current_inventory - monthly_demands[month_idx]
            monthly_inventory_levels[month_idx] = current_inventory

        return monthly_orders, monthly_inventory_levels

    def read_inventory_data(self, file_path):
        """Read the inventory Excel file and parse the data structure"""
        try:
            df = pd.read_excel(file_path, sheet_name=0)
            self.log(f"Successfully read {len(df)} rows from {file_path}")
        except Exception as e:
            self.log(f"Failed to read Excel file: {e}", "ERROR")
            raise

        month_columns = ['NOV\'24', 'DEC\'24', 'JAN\'25', 'FEB\'25', 'MAR\'25',
                         'APR\'25', 'MAY\'25', 'JUN\'25', 'JUL\'25', 'AUG\'25',
                         'SEP\'25', 'OCT\'25']

        base_columns = ['Supplier', 'sanwa item', 'sanwa item code', 'RM', 'Color', 'Revised Leadtime']

        return df, month_columns, base_columns

    def process_inventory_file(self, input_file_path, output_file_path):
        """Main function to process the inventory file and generate ordering recommendations"""
        self.log(f"Processing inventory file: {input_file_path}")

        df, month_columns, base_columns = self.read_inventory_data(input_file_path)

        stock_col = self.find_stock_on_hand_column(df)
        po_col = self.find_po_balance_column(df)

        if stock_col is None:
            self.log("Warning: Could not find stock on hand column. Using default value of 0.")
            stock_col = 'dummy_stock_col'
            df[stock_col] = 0

        if po_col is None:
            self.log("Warning: Could not find PO balance column. Using default value of 0.")
            po_col = 'dummy_po_col'
            df[po_col] = 0

        self.log(f"Using stock column: {stock_col}")
        self.log(f"Using PO balance column: {po_col}")

        output_df = df[base_columns].copy()

        # Add original demand columns for reference
        for month in month_columns:
            if month in df.columns:
                output_df[f'Demand_{month}'] = df[month].apply(self.safe_int_convert)

        # Add starting inventory information
        if stock_col in df.columns:
            output_df['Starting_Stock_on_Hand'] = df[stock_col].apply(self.safe_int_convert)
        if po_col in df.columns:
            output_df['Starting_PO_Balance'] = df[po_col].apply(self.safe_int_convert)

        # Calculate orders for each row
        self.log("Calculating ordering recommendations...")
        for idx, row in df.iterrows():
            monthly_orders, monthly_inventory_levels = self.inventory_ordering_policy(row, month_columns, stock_col,
                                                                                      po_col)

            for i, month in enumerate(month_columns):
                order_col_name = f'Order_{month}'
                inventory_col_name = f'Inventory_{month}'

                output_df.loc[idx, order_col_name] = monthly_orders[i]
                output_df.loc[idx, inventory_col_name] = monthly_inventory_levels[i]

        # Calculate total orders
        order_cols = [col for col in output_df.columns if col.startswith('Order_')]
        output_df['Total_Orders'] = output_df[order_cols].sum(axis=1)

        # Save to Excel
        self.log(f"Saving results to: {output_file_path}")
        self.save_to_excel(output_df, output_file_path)

        return output_df

    def save_to_excel(self, output_df, output_file_path):
        """Save dataframe to Excel with formatting"""
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            output_df.to_excel(writer, sheet_name='Ordering_Recommendations', index=False)

            worksheet = writer.sheets['Ordering_Recommendations']

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 15)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Color coding
            from openpyxl.styles import PatternFill

            order_fill_first = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
            order_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            inventory_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")

            First_order_set = True

            for col_idx, column in enumerate(worksheet.columns, 1):
                col_letter = column[0].column_letter
                header_cell = worksheet[f"{col_letter}1"]

                if header_cell.value and isinstance(header_cell.value, str):
                    if header_cell.value.startswith('Order_'):
                        if First_order_set:
                            for row_idx in range(1, len(output_df) + 2):
                                worksheet[f"{col_letter}{row_idx}"].fill = order_fill_first
                            First_order_set = False
                        else:
                            for row_idx in range(1, len(output_df) + 2):
                                worksheet[f"{col_letter}{row_idx}"].fill = order_fill
                    elif header_cell.value.startswith('Inventory_'):
                        for row_idx in range(1, len(output_df) + 2):
                            worksheet[f"{col_letter}{row_idx}"].fill = inventory_fill

    def generate_summary_report(self, output_df, save_to_file=None):
        """Generate a summary report of the ordering recommendations"""
        report_lines = []
        report_lines.append("=== INVENTORY ORDERING SUMMARY REPORT ===")
        report_lines.append(f"Total items processed: {len(output_df)}")
        report_lines.append(f"Items requiring orders: {len(output_df[output_df['Total_Orders'] > 0])}")
        report_lines.append(f"Total order value: {output_df['Total_Orders'].sum():.2f} kg")
        report_lines.append("")

        # Top items by total order quantity
        report_lines.append("TOP 10 ITEMS BY ORDER QUANTITY:")
        top_items = output_df.nlargest(10, 'Total_Orders')[
            ['Supplier', 'sanwa item code', 'RM', 'Color', 'Total_Orders']]
        report_lines.append(top_items.to_string(index=False))
        report_lines.append("")

        # Monthly order summary
        order_cols = [col for col in output_df.columns if col.startswith('Order_')]
        inventory_cols = [col for col in output_df.columns if col.startswith('Inventory_')]

        monthly_order_totals = output_df[order_cols].sum()
        monthly_avg_inventory = output_df[inventory_cols].mean()

        report_lines.append("MONTHLY ORDER TOTALS:")
        for col, total in monthly_order_totals.items():
            month = col.replace('Order_', '')
            report_lines.append(f"{month}: {total:.2f} kg")

        report_lines.append("\nAVERAGE MONTHLY INVENTORY LEVELS:")
        for col, avg_inv in monthly_avg_inventory.items():
            month = col.replace('Inventory_', '')
            report_lines.append(f"{month}: {avg_inv:.2f} kg")

        # Low inventory items
        report_lines.append("\nITEMS WITH LOW ENDING INVENTORY (<100kg in last month):")
        last_inventory_col = inventory_cols[-1] if inventory_cols else None
        if last_inventory_col:
            low_inventory_items = output_df[output_df[last_inventory_col] < 100]
            if len(low_inventory_items) > 0:
                low_inv_display = low_inventory_items[
                    ['Supplier', 'sanwa item code', 'RM', 'Color', last_inventory_col]].head(10)
                report_lines.append(low_inv_display.to_string(index=False))
            else:
                report_lines.append("No items with critically low inventory levels.")

        report_text = "\n".join(report_lines)

        if save_to_file:
            with open(save_to_file, 'w') as f:
                f.write(report_text)
            self.log(f"Summary report saved to: {save_to_file}")

        print(report_text)
        return report_text


def create_parser():
    """Create command line argument parser"""
    parser = argparse.ArgumentParser(
        description="Inventory Management CLI Tool - Process inventory data and generate ordering recommendations",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s                          # Process inventory file in current directory
  %(prog)s -i inventory.xlsx        # Process specific file
  %(prog)s -i data/ -o results/     # Specify input and output directories
  %(prog)s --verbose                # Enable verbose output
  %(prog)s --version                # Show version info
        """
    )

    parser.add_argument('-i', '--input',
                        help='Input file or directory (default: current directory)')

    parser.add_argument('-o', '--output',
                        help='Output directory (default: current directory)')

    parser.add_argument('-v', '--verbose',
                        action='store_true',
                        help='Enable verbose output')

    parser.add_argument('--version',
                        action='version',
                        version=f'%(prog)s {__version__}')

    parser.add_argument('--no-report',
                        action='store_true',
                        help='Skip generating summary report')

    parser.add_argument('--save-report',
                        help='Save summary report to specified file')

    return parser


def main():
    """Main CLI entry point"""
    parser = create_parser()
    args = parser.parse_args()

    # Initialize inventory manager
    manager = InventoryManager(verbose=args.verbose)

    try:
        # Determine input file/directory
        if args.input:
            if os.path.isfile(args.input):
                input_file = args.input
            elif os.path.isdir(args.input):
                input_file = manager.find_input_file(args.input)
                if not input_file:
                    manager.log(f"No inventory Excel file found in directory: {args.input}", "ERROR")
                    return 1
            else:
                manager.log(f"Input path does not exist: {args.input}", "ERROR")
                return 1
        else:
            # Default: look in current directory
            input_file = manager.find_input_file(".")
            if not input_file:
                manager.log("No inventory Excel file found in current directory", "ERROR")
                manager.log("Please ensure your inventory Excel file is in the current directory", "ERROR")
                return 1

        manager.log(f"Found input file: {input_file}")

        # Determine output directory
        output_dir = args.output if args.output else "."
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            manager.log(f"Created output directory: {output_dir}")

        # Generate output filename
        output_file = manager.generate_output_filename(input_file, output_dir)
        manager.log(f"Output will be saved as: {output_file}")

        # Process the inventory file
        result_df = manager.process_inventory_file(input_file, output_file)

        # Generate summary report
        if not args.no_report:
            report_file = args.save_report if args.save_report else None
            manager.generate_summary_report(result_df, report_file)

        manager.log(f"SUCCESS: Ordering recommendations saved to: {output_file}", "SUCCESS")

        return 0

    except FileNotFoundError as e:
        manager.log(f"File not found: {e}", "ERROR")
        return 1
    except Exception as e:
        manager.log(f"Failed to process file: {str(e)}", "ERROR")
        if args.verbose:
            import traceback
            traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())