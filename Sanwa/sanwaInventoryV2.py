import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
import re

def read_inventory_data(file_path):
    """
    Read the inventory Excel file and parse the data structure
    """
    # Read the Excel file
    df = pd.read_excel(file_path, sheet_name=0)

    # Define the month columns (adjust based on your actual column names)
    month_columns = ['NOV\'24', 'DEC\'24', 'JAN\'25', 'FEB\'25', 'MAR\'25',
                     'APR\'25', 'MAY\'25', 'JUN\'25', 'JUL\'25', 'AUG\'25',
                     'SEP\'25', 'OCT\'25']

    # First 5 columns to keep unchanged
    base_columns = ['Supplier', 'sanwa item', 'sanwa item code', 'RM', 'Color', 'Revised Leadtime']

    return df, month_columns, base_columns


def parse_leadtime(leadtime_str):
    """
    Parse leadtime string and return weeks as integer
    Examples:
    "10-11wks MOQ1MT" -> 11 (taking the upper bound)
    "24wks,germany MOQ 10,000kg" -> 24
    "4mth w/fcst" -> 16 (4 months * 4 weeks)
    "22-30wks/EUR" -> 30
    "20wks / EOL 30/10/23" -> 20
    """
    if pd.isna(leadtime_str):
        return 8  # Default leadtime

    leadtime_str = str(leadtime_str).strip()

    # Handle months first (convert to weeks)
    month_pattern = r'(\d+)\s*mth'
    month_match = re.search(month_pattern, leadtime_str, re.IGNORECASE)
    if month_match:
        months = int(month_match.group(1))
        return months * 4  # Convert months to weeks (4 weeks per month)

    # Look for week patterns with various formats
    # Pattern 1: Range like "X-Ywks", "X-Ywk", "X-Y wks"
    week_range_pattern = r'(\d+)\s*[-–]\s*(\d+)\s*wks?'
    match = re.search(week_range_pattern, leadtime_str, re.IGNORECASE)
    if match:
        start_weeks = int(match.group(1))
        end_weeks = int(match.group(2))
        return max(start_weeks, end_weeks)  # Take the upper bound

    # Pattern 2: Single week number like "20wks", "24wks"
    single_week_pattern = r'(\d+)\s*wks?'
    match = re.search(single_week_pattern, leadtime_str, re.IGNORECASE)
    if match:
        return int(match.group(1))

    # Pattern 3: Just numbers that could be weeks (but filter out obvious MOQ values)
    all_numbers = re.findall(r'\d+', leadtime_str)
    if all_numbers:
        # Filter out numbers that are likely MOQ (very large or have kg/MT context)
        leadtime_candidates = []
        for num_str in all_numbers:
            num = int(num_str)
            # Check if this number appears near MOQ, kg, MT indicators
            num_pos = leadtime_str.find(num_str)
            context = leadtime_str[max(0, num_pos - 10):num_pos + len(num_str) + 10].upper()

            # Skip if it's clearly MOQ related
            if any(indicator in context for indicator in ['MOQ', 'KG', 'MT']):
                continue

            # Accept reasonable leadtime values (typically 1-52 weeks)
            if 1 <= num <= 52:
                leadtime_candidates.append(num)

        if leadtime_candidates:
            return max(leadtime_candidates)

    return 8  # Default


def extract_moq(leadtime_str):
    """
    Extract MOQ (Minimum Order Quantity) from leadtime string
    Examples:
    "10-11wks MOQ1MT" -> 1000 (1MT = 1000kg)
    "24wks,germany MOQ 10,000kg" -> 10000
    "20wks,MOQ500kg" -> 500
    "MOQ: 5MT 16-20wks" -> 5000
    "20wks/MOQ3MT" -> 3000
    "MOQ: 1mt 10-11wks" -> 1000 (case insensitive)
    """
    if pd.isna(leadtime_str):
        return 0

    leadtime_str = str(leadtime_str).strip()
    import re

    # Pattern 1: MOQ with MT/mt (metric tons) - handle various formats
    # Matches: "MOQ: 1MT", "MOQ1MT", "MOQ 5MT", "MOQ: 5mt", "/MOQ3MT"
    moq_mt_patterns = [
        r'MOQ\s*:?\s*(\d+(?:,\d+)*(?:\.\d+)?)\s*MT',  # MOQ: 1MT, MOQ 1MT
        r'/MOQ(\d+(?:,\d+)*(?:\.\d+)?)MT',  # /MOQ3MT
        r'MOQ(\d+(?:,\d+)*(?:\.\d+)?)MT'  # MOQ1MT
    ]

    for pattern in moq_mt_patterns:
        match = re.search(pattern, leadtime_str, re.IGNORECASE)
        if match:
            # Remove commas from number and convert
            number_str = match.group(1).replace(',', '')
            return int(float(number_str) * 1000)  # Convert MT to kg

    # Pattern 2: MOQ with kg - handle various formats and comma separators
    # Matches: "MOQ 10,000kg", "MOQ500kg", "MOQ: 5000kg", "MOQ 1000kg"
    moq_kg_patterns = [
        r'MOQ\s*:?\s*(\d+(?:,\d+)*(?:\.\d+)?)\s*kg',  # MOQ: 1000kg, MOQ 10,000kg
        r',MOQ(\d+(?:,\d+)*(?:\.\d+)?)kg',  # ,MOQ500kg
        r'/MOQ(\d+(?:,\d+)*(?:\.\d+)?)kg'  # /MOQ1000kg (from "20wks/MOQ3MT")
    ]

    for pattern in moq_kg_patterns:
        match = re.search(pattern, leadtime_str, re.IGNORECASE)
        if match:
            # Remove commas from number
            number_str = match.group(1).replace(',', '')
            return int(float(number_str))

    # Pattern 3: MOQ with just numbers (assume kg unless very small)
    # Matches: "MOQ 5000", "MOQ: 1000", "MOQ 300"
    moq_number_patterns = [
        r'MOQ\s*:?\s*(\d+(?:,\d+)*)',  # MOQ: 5000, MOQ 1000
        r'MOQ(\d+(?:,\d+)*)'  # MOQ5000
    ]

    for pattern in moq_number_patterns:
        match = re.search(pattern, leadtime_str, re.IGNORECASE)
        if match:
            number_str = match.group(1).replace(',', '')
            number = int(number_str)

            # If it's a small number (≤10), assume it's metric tons
            if number <= 10:
                return number * 1000
            else:
                return number

    return 0  # No MOQ found
def calculate_safety_stock(demands, leadtime_weeks):
    """
    Calculate safety stock based on demand variability and leadtime
    """
    if len(demands) == 0 or all(pd.isna(demands)):
        return 0

    # Remove NaN values
    clean_demands = [d for d in demands if not pd.isna(d) and d > 0]
    if len(clean_demands) == 0:
        return 0

    # Calculate standard deviation of demand
    demand_std = np.std(clean_demands) if len(clean_demands) > 1 else np.mean(clean_demands) * 0.2

    # Safety stock = Z-score * demand_std * sqrt(leadtime)
    # Using Z-score of 1.65 for 95% service level
    safety_stock = 1.65 * demand_std * np.sqrt(leadtime_weeks / 4.33)  # Convert weeks to months

    return max(safety_stock, 0)


def inventory_ordering_policy(row, month_columns):
    """
    Implement inventory ordering policy for a single item
    Returns: (monthly_orders, monthly_inventory_levels)
    """
    # Extract basic information
    leadtime_weeks = parse_leadtime(row.get('Revised Leadtime'))
    stock_on_hand = row.get('Sanwa stock on hand 27/11/2024', 0)
    po_balance = row.get('Sanwa system PO bal as at ', 0)

    # Handle NaN values
    stock_on_hand = 0 if pd.isna(stock_on_hand) else float(stock_on_hand)
    po_balance = 0 if pd.isna(po_balance) else float(po_balance)

    # Extract monthly demands
    monthly_demands = []
    for month in month_columns:
        demand = row.get(month, 0)
        demand = 0 if pd.isna(demand) else float(demand)
        monthly_demands.append(demand)

    # Calculate safety stock
    safety_stock = calculate_safety_stock(monthly_demands, leadtime_weeks)

    # Initialize variables
    num_months = len(month_columns)
    leadtime_months = max(1, round(leadtime_weeks / 4.33))  # Convert weeks to months

    # Extract MOQ from leadtime information
    leadtime_info = str(row.get('Revised Leadtime', ''))
    moq = extract_moq(leadtime_info)

    # Initialize arrays to track orders, deliveries, and inventory
    monthly_orders = [0.0] * num_months
    monthly_deliveries = [0.0] * num_months
    monthly_inventory_levels = [0.0] * num_months

    # Set up initial conditions
    current_inventory = stock_on_hand

    # Assume existing PO balance arrives in the first month (or distribute over first few months)
    if po_balance > 0:
        # Assume existing PO arrives in month 0 (could be adjusted based on your business logic)
        monthly_deliveries[0] += po_balance

    # Calculate month by month
    for month_idx in range(num_months):
        # Start of month: add any deliveries from previous orders
        current_inventory += monthly_deliveries[month_idx]

        # Calculate projected inventory needs for decision making
        # Look ahead to determine if we need to order
        future_demand = 0
        look_ahead_months = min(leadtime_months, num_months - month_idx)
        for j in range(look_ahead_months):
            if month_idx + j < num_months:
                future_demand += monthly_demands[month_idx + j]

        reorder_point = future_demand + safety_stock

        # Check if we need to order (before consuming this month's demand)
        projected_inventory_after_demand = current_inventory - monthly_demands[month_idx]

        if projected_inventory_after_demand < reorder_point:
            # Calculate order quantity
            order_qty = reorder_point - projected_inventory_after_demand

            # Apply MOQ constraint
            if moq > 0 and order_qty > 0:
                order_qty = max(order_qty, moq)
                # Round up to nearest MOQ multiple if needed
                if order_qty > moq:
                    order_qty = np.ceil(order_qty / moq) * moq

            monthly_orders[month_idx] = order_qty

            # Schedule delivery based on lead time
            delivery_month = month_idx + leadtime_months
            if delivery_month < num_months:
                monthly_deliveries[delivery_month] += order_qty

        # Consume this month's demand
        current_inventory = max(0, current_inventory - monthly_demands[month_idx])

        # Record end-of-month inventory level
        monthly_inventory_levels[month_idx] = current_inventory

    return monthly_orders, monthly_inventory_levels


def process_inventory_file(input_file_path, output_file_path):
    """
    Main function to process the inventory file and generate ordering recommendations
    """
    # Read the data
    df, month_columns, base_columns = read_inventory_data(input_file_path)

    # Create output dataframe with base columns
    output_df = df[base_columns].copy()

    # Add original demand columns for reference
    for month in month_columns:
        if month in df.columns:
            output_df[f'Demand_{month}'] = df[month]

    # Add starting inventory information for reference
    if 'Sanwa stock on hand 27/11/2024' in df.columns:
        output_df['Starting_Stock_on_Hand'] = df['Sanwa stock on hand 27/11/2024']
    if 'Sanwa system PO bal as at ' in df.columns:
        output_df['Starting_PO_Balance'] = df['Sanwa system PO bal as at ']

    # Calculate orders for each row
    order_columns = [f'Order_{month}' for month in month_columns]

    for idx, row in df.iterrows():
        monthly_orders, monthly_inventory_levels = inventory_ordering_policy(row, month_columns)

        # Add orders and inventory levels to output dataframe
        for i, month in enumerate(month_columns):
            order_col_name = f'Order_{month}'
            inventory_col_name = f'Inventory_{month}'

            output_df.loc[idx, order_col_name] = monthly_orders[i]
            output_df.loc[idx, inventory_col_name] = monthly_inventory_levels[i]

    # Calculate total orders
    order_cols = [col for col in output_df.columns if col.startswith('Order_')]
    output_df['Total_Orders'] = output_df[order_cols].sum(axis=1)

    # Save to Excel
    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        output_df.to_excel(writer, sheet_name='Ordering_Recommendations', index=False)

        # Format the Excel file
        worksheet = writer.sheets['Ordering_Recommendations']

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 15)  # Reduced width for better fit
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Color code order and inventory columns for easier reading
        from openpyxl.styles import PatternFill

        # Light blue for order columns
        order_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        # Light green for inventory columns
        inventory_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")

        # Apply colors to headers and data
        for col_idx, column in enumerate(worksheet.columns, 1):
            col_letter = column[0].column_letter
            header_cell = worksheet[f"{col_letter}1"]

            if header_cell.value and isinstance(header_cell.value, str):
                if header_cell.value.startswith('Order_'):
                    # Color entire order column
                    for row_idx in range(1, len(output_df) + 2):
                        worksheet[f"{col_letter}{row_idx}"].fill = order_fill
                elif header_cell.value.startswith('Inventory_'):
                    # Color entire inventory column
                    for row_idx in range(1, len(output_df) + 2):
                        worksheet[f"{col_letter}{row_idx}"].fill = inventory_fill

    return output_df
def generate_summary_report(output_df):
    """
    Generate a summary report of the ordering recommendations
    """
    print("=== INVENTORY ORDERING SUMMARY REPORT ===")
    print(f"Total items processed: {len(output_df)}")
    print(f"Items requiring orders: {len(output_df[output_df['Total_Orders'] > 0])}")
    print(f"Total order value: {output_df['Total_Orders'].sum():.2f} kg")
    print()

    # Show top items by total order quantity
    print("TOP 10 ITEMS BY ORDER QUANTITY:")
    top_items = output_df.nlargest(10, 'Total_Orders')[['Supplier', 'sanwa item code', 'RM', 'Color', 'Total_Orders']]
    print(top_items.to_string(index=False))
    print()

    # Monthly order summary
    order_cols = [col for col in output_df.columns if col.startswith('Order_')]
    inventory_cols = [col for col in output_df.columns if col.startswith('Inventory_')]

    monthly_order_totals = output_df[order_cols].sum()
    monthly_avg_inventory = output_df[inventory_cols].mean()

    print("MONTHLY ORDER TOTALS:")
    for col, total in monthly_order_totals.items():
        month = col.replace('Order_', '')
        print(f"{month}: {total:.2f} kg")

    print("\nAVERAGE MONTHLY INVENTORY LEVELS:")
    for col, avg_inv in monthly_avg_inventory.items():
        month = col.replace('Inventory_', '')
        print(f"{month}: {avg_inv:.2f} kg")

    # Show items with potential stockout risk (low ending inventory)
    print("\nITEMS WITH LOW ENDING INVENTORY (<100kg in last month):")
    last_inventory_col = inventory_cols[-1] if inventory_cols else None
    if last_inventory_col:
        low_inventory_items = output_df[output_df[last_inventory_col] < 100]
        if len(low_inventory_items) > 0:
            low_inv_display = low_inventory_items[['Supplier', 'sanwa item code', 'RM', 'Color', last_inventory_col]].head(
                10)
            print(low_inv_display.to_string(index=False))
        else:
            print("No items with critically low inventory levels.")


# Example usage
if __name__ == "__main__":
    # Replace with your actual file paths
    input_file = "C:/Users/I3Nexus/Desktop/inventory_data.xlsx"
    output_file = "C:/Users/I3Nexus/Desktop/ordering_recommendations_V2.xlsx"

    try:
        # Process the inventory file
        result_df = process_inventory_file(input_file, output_file)

        # Generate summary report
        generate_summary_report(result_df)

        print(f"\nOrdering recommendations saved to: {output_file}")

    except FileNotFoundError:
        print(f"Error: Could not find the input file '{input_file}'")
        print("Please make sure the file exists and the path is correct.")
    except Exception as e:
        print(f"Error processing file: {str(e)}")
        print("Please check your data format and try again.")